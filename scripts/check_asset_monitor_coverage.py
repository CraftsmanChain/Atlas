#!/usr/bin/env python3
"""Compare asset nodes with Prometheus or VM query API results."""

from __future__ import annotations

import argparse
import csv
import json
import os
import urllib.parse
import urllib.request
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_JOBS = ("node_exporter", "dcgm_exporter", "gpu_exporter", "ipmi_exporter")
DEFAULT_GPU_JOBS = ("dcgm_exporter", "gpu_exporter", "ipmi_exporter")


@dataclass
class AssetNode:
    ip: str
    bmc_ip: str
    name: str
    power_status: str
    agent_status: str
    verify_status: str
    tags: str
    config_info: str


def normalize_ip(raw: object) -> str:
    return str(raw or "").strip()


def instance_ip(instance: str) -> str:
    value = str(instance or "").strip()
    if value.count(":") == 1 and "." in value:
        return value.rsplit(":", 1)[0]
    return value


def split_csv_arg(value: str) -> tuple[str, ...]:
    return tuple(item.strip() for item in value.split(",") if item.strip())


def write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def derive_bmc_ip(host_ip: str, explicit_bmc_ip: str, host_prefix: str, bmc_prefix: str) -> str:
    if explicit_bmc_ip:
        return explicit_bmc_ip
    if host_prefix and bmc_prefix and host_ip.startswith(host_prefix):
        last_octet = host_ip.split(".")[-1]
        return f"{bmc_prefix}{last_octet}"
    return ""


def load_assets_from_xlsx(
    path: Path,
    ip_column: str,
    name_column: str,
    bmc_column: str,
    ipmi_host_prefix: str,
    ipmi_bmc_prefix: str,
) -> list[AssetNode]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [str(value).strip() if value is not None else "" for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    header_index = {name: index for index, name in enumerate(headers)}
    if ip_column not in header_index:
        raise ValueError(f"资产文件缺少 IP 列: {ip_column}")
    if name_column not in header_index:
        raise ValueError(f"资产文件缺少名称列: {name_column}")

    def get(row: tuple[object, ...], column: str) -> str:
        index = header_index.get(column)
        return str(row[index] or "").strip() if index is not None and index < len(row) else ""

    rows: list[AssetNode] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        ip = normalize_ip(values[header_index[ip_column]])
        if not ip:
            continue
        rows.append(
            AssetNode(
                ip=ip,
                bmc_ip=derive_bmc_ip(ip, get(values, bmc_column), ipmi_host_prefix, ipmi_bmc_prefix),
                name=get(values, name_column),
                power_status=get(values, "状态"),
                agent_status=get(values, "监控 Agent"),
                verify_status=get(values, "验证状态"),
                tags=get(values, "标签"),
                config_info=get(values, "配置信息"),
            )
        )
    return rows


def load_assets_from_csv(
    path: Path,
    ip_column: str,
    name_column: str,
    bmc_column: str,
    ipmi_host_prefix: str,
    ipmi_bmc_prefix: str,
) -> list[AssetNode]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if ip_column not in (reader.fieldnames or []):
            raise ValueError(f"资产文件缺少 IP 列: {ip_column}")
        if name_column not in (reader.fieldnames or []):
            raise ValueError(f"资产文件缺少名称列: {name_column}")
        rows = []
        for row in reader:
            ip = normalize_ip(row.get(ip_column))
            if not ip:
                continue
            rows.append(
                AssetNode(
                    ip=ip,
                    bmc_ip=derive_bmc_ip(ip, str(row.get(bmc_column, "") or "").strip(), ipmi_host_prefix, ipmi_bmc_prefix),
                    name=str(row.get(name_column, "") or "").strip(),
                    power_status=str(row.get("状态", "") or "").strip(),
                    agent_status=str(row.get("监控 Agent", "") or "").strip(),
                    verify_status=str(row.get("验证状态", "") or "").strip(),
                    tags=str(row.get("标签", "") or "").strip(),
                    config_info=str(row.get("配置信息", "") or "").strip(),
                )
            )
        return rows


def load_asset_nodes(
    path: Path,
    ip_column: str,
    name_column: str,
    bmc_column: str,
    ipmi_host_prefix: str,
    ipmi_bmc_prefix: str,
) -> list[AssetNode]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return load_assets_from_xlsx(path, ip_column, name_column, bmc_column, ipmi_host_prefix, ipmi_bmc_prefix)
    if suffix == ".csv":
        return load_assets_from_csv(path, ip_column, name_column, bmc_column, ipmi_host_prefix, ipmi_bmc_prefix)
    raise ValueError(f"暂不支持的资产文件类型: {path.suffix}")


def query_api(query_api_url: str, promql: str) -> list[dict]:
    url = f"{query_api_url}?{urllib.parse.urlencode({'query': promql})}"
    with urllib.request.urlopen(url, timeout=60) as response:
        payload = json.load(response)
    if payload.get("status") != "success":
        raise RuntimeError(payload)
    return payload.get("data", {}).get("result", [])


def load_job_series(query_api_url: str, jobs: tuple[str, ...]) -> dict[str, dict[str, str]]:
    job_to_ip_status: dict[str, dict[str, str]] = {job: {} for job in jobs}
    for job in jobs:
        for row in query_api(query_api_url, f'up{{job="{job}"}}'):
            metric = row.get("metric", {})
            instance = metric.get("instance", "")
            ip = instance_ip(instance)
            if not ip:
                continue
            value = row.get("value", [])
            raw = value[1] if len(value) >= 2 else ""
            try:
                status = "UP" if float(raw) == 1 else "DOWN"
            except (TypeError, ValueError):
                status = "UNKNOWN"
            # 同一 job 同一 IP 若有多个 target，只要有一个 UP 就视为 UP。
            previous = job_to_ip_status[job].get(ip)
            if previous != "UP":
                job_to_ip_status[job][ip] = status
    return job_to_ip_status


def infer_gpu_candidate(ip: str, name: str, config_info: str, jobs_by_ip: dict[str, set[str]], gpu_jobs: tuple[str, ...]) -> bool:
    if any(job in jobs_by_ip.get(ip, set()) for job in gpu_jobs):
        return True
    text = f"{name} {config_info}".lower()
    return "gpu" in text or "nvidia" in text


def status_for(job_to_ip_status: dict[str, dict[str, str]], job: str, ip: str) -> str:
    return job_to_ip_status.get(job, {}).get(ip, "MISSING")


def target_ip_for_job(node: AssetNode, job: str) -> str:
    if job == "ipmi_exporter":
        return node.bmc_ip or ""
    return node.ip


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--asset-file", default=os.getenv("ASSET_FILE_PATH"), help="资产文件路径，支持 xlsx/csv")
    parser.add_argument("--query-api-url", default=os.getenv("QUERY_API_URL"), help="Prometheus/VM 查询接口，例如 http://host:9090/api/v1/query")
    parser.add_argument("--output-dir", default=os.getenv("OUTPUT_DIR", "docs/assets"))
    parser.add_argument("--asset-ip-column", default=os.getenv("ASSET_IP_COLUMN", "主机IP地址"))
    parser.add_argument("--asset-name-column", default=os.getenv("ASSET_NAME_COLUMN", "名称"))
    parser.add_argument("--asset-bmc-column", default=os.getenv("ASSET_BMC_COLUMN", ""))
    parser.add_argument("--jobs", default=os.getenv("MONITOR_JOBS", ",".join(DEFAULT_JOBS)))
    parser.add_argument("--gpu-jobs", default=os.getenv("GPU_JOBS", ",".join(DEFAULT_GPU_JOBS)))
    parser.add_argument("--ipmi-host-prefix", default=os.getenv("IPMI_HOST_PREFIX", "10.114.4."))
    parser.add_argument("--ipmi-bmc-prefix", default=os.getenv("IPMI_BMC_PREFIX", "10.114.1."))
    args = parser.parse_args()

    if not args.asset_file:
        raise ValueError("必须提供 --asset-file 或环境变量 ASSET_FILE_PATH")
    if not args.query_api_url:
        raise ValueError("必须提供 --query-api-url 或环境变量 QUERY_API_URL")

    asset_file = Path(args.asset_file)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    jobs = split_csv_arg(args.jobs)
    gpu_jobs = split_csv_arg(args.gpu_jobs)
    asset_nodes = load_asset_nodes(
        asset_file,
        args.asset_ip_column,
        args.asset_name_column,
        args.asset_bmc_column,
        args.ipmi_host_prefix,
        args.ipmi_bmc_prefix,
    )
    job_to_ip_status = load_job_series(args.query_api_url, jobs)

    jobs_by_ip: dict[str, set[str]] = {}
    for job, ip_map in job_to_ip_status.items():
        for ip in ip_map:
            jobs_by_ip.setdefault(ip, set()).add(job)

    asset_ip_to_node = {node.ip: node for node in asset_nodes}
    asset_ips = set(asset_ip_to_node)
    expected_targets: dict[tuple[str, str], str] = {}
    covered_monitor_targets: set[tuple[str, str]] = set()
    for node in asset_nodes:
        gpu_candidate = infer_gpu_candidate(node.ip, node.name, node.config_info, jobs_by_ip, gpu_jobs)
        for job in jobs:
            if job in gpu_jobs and not gpu_candidate:
                continue
            target_ip = target_ip_for_job(node, job)
            if target_ip:
                expected_targets[(job, target_ip)] = node.ip

    collected_at = datetime.now(timezone.utc).astimezone()
    stamp = collected_at.strftime("%Y-%m-%d")

    compare_rows: list[dict[str, str]] = []
    asset_only_rows: list[dict[str, str]] = []
    monitor_only_rows: list[dict[str, str]] = []
    non_up_rows: list[dict[str, str]] = []
    issue_counts: Counter[str] = Counter()

    for ip in sorted(asset_ips, key=lambda value: tuple(int(part) for part in value.split("."))):
        node = asset_ip_to_node[ip]
        asset_name = node.name
        exists_in_asset = True
        gpu_candidate = infer_gpu_candidate(ip, asset_name, node.config_info, jobs_by_ip, gpu_jobs)

        row: dict[str, str] = {
            "node_ip": ip,
            "bmc_ip": node.bmc_ip,
            "name": asset_name,
            "in_asset": "YES",
            "in_monitor": "NO",
            "gpu_candidate": "YES" if gpu_candidate else "NO",
            "platform_power_status": node.power_status,
            "platform_agent_status": node.agent_status,
            "platform_verify_status": node.verify_status,
            "tags": node.tags,
        }

        has_monitor = False
        has_non_up = False
        for job in jobs:
            if job in gpu_jobs and not gpu_candidate:
                row[job] = "N/A"
                continue
            target_ip = target_ip_for_job(node, job)
            if not target_ip:
                row[job] = "MISSING"
                has_non_up = True
                issue_counts[f"{job}_MISSING"] += 1
                continue
            status = status_for(job_to_ip_status, job, target_ip)
            row[job] = status
            if status != "MISSING":
                has_monitor = True
                covered_monitor_targets.add((job, target_ip))
            if status != "UP":
                has_non_up = True
                issue_counts[f"{job}_{status}"] += 1

        row["in_monitor"] = "YES" if has_monitor else "NO"
        compare_rows.append(row)

        if has_non_up:
            non_up_rows.append(
                {
                    "node_ip": ip,
                    "bmc_ip": node.bmc_ip,
                    "name": asset_name,
                    "gpu_candidate": row["gpu_candidate"],
                    "platform_power_status": node.power_status,
                    "platform_agent_status": node.agent_status,
                    "platform_verify_status": node.verify_status,
                    **{job: row.get(job, "N/A") for job in jobs},
                }
            )

        if not has_monitor:
            asset_only_rows.append(
                {
                    "node_ip": ip,
                    "name": asset_name,
                    "bmc_ip": node.bmc_ip,
                    "platform_power_status": node.power_status,
                    "platform_agent_status": node.agent_status,
                    "platform_verify_status": node.verify_status,
                    "tags": node.tags,
                }
            )
            issue_counts["asset_has_but_monitor_missing"] += 1

    extra_targets: dict[str, set[str]] = {}
    for job, ip_map in job_to_ip_status.items():
        for target_ip in ip_map:
            if (job, target_ip) in covered_monitor_targets:
                continue
            extra_targets.setdefault(target_ip, set()).add(job)

    for target_ip in sorted(extra_targets, key=lambda value: tuple(int(part) for part in value.split("."))):
        monitor_only_rows.append(
            {
                "target_ip": target_ip,
                "jobs": ",".join(sorted(extra_targets[target_ip])),
                "mapped_asset_node_ip": expected_targets.get((next(iter(extra_targets[target_ip])), target_ip), ""),
            }
        )
        issue_counts["monitor_has_but_asset_missing"] += 1

    compare_columns = [
        "node_ip",
        "bmc_ip",
        "name",
        "in_asset",
        "in_monitor",
        "gpu_candidate",
        "platform_power_status",
        "platform_agent_status",
        "platform_verify_status",
        "tags",
        *jobs,
    ]

    compare_csv = output_dir / f"asset-monitor-compare-{stamp}.csv"
    asset_only_csv = output_dir / f"asset-has-monitor-missing-{stamp}.csv"
    monitor_only_csv = output_dir / f"monitor-has-asset-missing-{stamp}.csv"
    non_up_csv = output_dir / f"monitor-status-not-up-{stamp}.csv"
    report_md = output_dir / f"asset-monitor-report-{stamp}.md"

    write_csv(compare_csv, compare_rows, compare_columns)
    write_csv(
        asset_only_csv,
        asset_only_rows,
        ["node_ip", "bmc_ip", "name", "platform_power_status", "platform_agent_status", "platform_verify_status", "tags"],
    )
    write_csv(monitor_only_csv, monitor_only_rows, ["target_ip", "jobs", "mapped_asset_node_ip"])
    write_csv(
        non_up_csv,
        non_up_rows,
        [
            "node_ip",
            "bmc_ip",
            "name",
            "gpu_candidate",
            "platform_power_status",
            "platform_agent_status",
            "platform_verify_status",
            *jobs,
        ],
    )

    with report_md.open("w", encoding="utf-8") as handle:
        handle.write(f"# 资产与监控覆盖核查（{stamp}）\n\n")
        handle.write(f"- 采集时间：{collected_at.isoformat(timespec='seconds')}\n")
        handle.write(f"- 资产文件：`{asset_file}`\n")
        handle.write(f"- 查询接口：`{args.query_api_url}`\n")
        handle.write(f"- 检查 jobs：`{','.join(jobs)}`\n")
        handle.write(f"- 资产节点数：{len(asset_ips)}\n")
        handle.write(f"- 监控目标 IP 数：{sum(len(ip_map) for ip_map in job_to_ip_status.values())}\n")
        handle.write(f"- 资产文件有、监控没有：{len(asset_only_rows)}\n")
        handle.write(f"- 监控有、资产文件没有：{len(monitor_only_rows)}\n")
        handle.write(f"- 状态非 UP 记录数：{len(non_up_rows)}\n\n")

        handle.write("## 问题统计\n\n| 类型 | 数量 |\n|---|---:|\n")
        for issue_type, count in sorted(issue_counts.items()):
            handle.write(f"| {issue_type} | {count} |\n")

        handle.write("\n## 资产文件有、监控没有\n\n| 节点 | BMC | 名称 | 平台状态 | 监控 Agent |\n|---|---|---|---|---|\n")
        for row in asset_only_rows:
            handle.write(f"| {row['node_ip']} | {row['bmc_ip']} | {row['name']} | {row['platform_power_status']} | {row['platform_agent_status']} |\n")

        handle.write("\n## 监控有、资产文件没有\n\n| 目标 IP | jobs | 映射资产节点 |\n|---|---|---|\n")
        for row in monitor_only_rows:
            handle.write(f"| {row['target_ip']} | {row['jobs']} | {row['mapped_asset_node_ip']} |\n")

        display_names = {
            "node_exporter": "node",
            "dcgm_exporter": "dcgm",
            "gpu_exporter": "gpu",
            "ipmi_exporter": "ipmi",
        }
        job_headers = " | ".join(display_names.get(job, job) for job in jobs)
        job_separators = "|".join("---" for _ in jobs)
        handle.write(
            f"\n## 状态不是 UP 的资产节点\n\n| 节点 | BMC | 名称 | {job_headers} |\n"
            f"|---|---|---|{job_separators}|\n"
        )
        for row in non_up_rows:
            statuses = " | ".join(row.get(job, "N/A") for job in jobs)
            handle.write(f"| {row['node_ip']} | {row['bmc_ip']} | {row['name']} | {statuses} |\n")

        handle.write("\n## 文件\n\n")
        handle.write(f"- `{compare_csv.name}`：全量对比\n")
        handle.write(f"- `{asset_only_csv.name}`：资产文件有、监控没有\n")
        handle.write(f"- `{monitor_only_csv.name}`：监控有、资产文件没有\n")
        handle.write(f"- `{non_up_csv.name}`：状态不是 UP 的节点\n")
        handle.write(f"- `{report_md.name}`：摘要报告\n")

    print(
        json.dumps(
            {
                "asset_nodes": len(asset_ips),
                "monitor_target_ips": sum(len(ip_map) for ip_map in job_to_ip_status.values()),
                "asset_has_but_monitor_missing": len(asset_only_rows),
                "monitor_has_but_asset_missing": len(monitor_only_rows),
                "status_not_up": len(non_up_rows),
                "files": [
                    str(compare_csv),
                    str(asset_only_csv),
                    str(monitor_only_csv),
                    str(non_up_csv),
                    str(report_md),
                ],
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
